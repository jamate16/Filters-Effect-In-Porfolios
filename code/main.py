from dataclasses import dataclass
from abc import ABC, abstractmethod
from enum import Enum

import openpyxl as opx
from openpyxl.utils import get_column_letter, column_index_from_string, coordinate_to_tuple

import os
import datetime
import numpy as np
import pandas as pd

from tqdm import tqdm

class FrequencyOfData(Enum):
    ANNUAL = 1
    QUARTERLY = 2

@dataclass
class FileStyleDetails:
    """
    Dataclass that stores the locations, names, and any other attribute that characterizes a sheet style
    
    Attributes:
        company_name_cell (str): The cell coordinate containing the company name.
        metric_sheet_name (str, optional): The name of the sheet where the metric is located (default is an empty string).
        metric_row_name (str, optional): The substring to search in the index of the rows of metric_sheet_name (default is an empty string).
        sheet_name_base (str): For comparison with sheet_name_compare. Cell coord where the sheet name is located.
        sheet_name_compare (str): For comparison with sheet_name_base. Cell coord where the sheet name is repeated.
    """
    company_name_cell: str
    date_format: str
    metric_sheet_name: str
    metric_row: int
    sheet_name_base: str
    sheet_name_compare: str
    metric_timestamp_seed: str

class FileStyle(Enum):
    A = "StyleA"
    B = "StyleB"

class FileStyleManager:
    def __init__(self, styles : dict):
        self.styles = styles

    def determine_file_style(self, workbook : opx.Workbook) -> FileStyle:
        """
        Compares the contents of two cells (different cells for every file style) that contain the same substring

        Returns:
            FileStyle: If the comparison for the corresponding file style returns true
        """
        active_sheet = workbook.active
        
        for style_name, style in self.styles.items():
            sheet_name_base = active_sheet[style.sheet_name_base].value
            sheet_name_compare = active_sheet[style.sheet_name_compare].value.split("-") #
            sheet_name_compare = sheet_name_compare[0].split("\xa0\xa0")[0].strip() # In style A there a bunch of "\xa0" characters

            if sheet_name_compare in sheet_name_base:
                return style_name
            
        raise Exception("Sheet style not recognized.")

class IMetricExtractor(ABC):
    def __init__(self, workbook : opx.Workbook, file_structure_details : FileStyleDetails):
        self.workbook = workbook
        self.file_structure_details = file_structure_details

        self.worksheet = None
        self.open_sheet()

        self.target_row = None

    def open_sheet(self) -> None:
        for sheet in self.workbook.worksheets:
            if self.file_structure_details.metric_sheet_name in sheet[self.file_structure_details.sheet_name_base].value:
                self.worksheet = sheet
                break

        if not self.worksheet:
            raise Exception(f"Metric sheet not found. {self.file_structure_details.metric_sheet_name} not in cell {self.file_structure_details.sheet_name_base} on any sheet")

    @abstractmethod
    def get_company_name(self) -> str:
        pass

    def get_metric_data(self) -> pd.Series:
        timestamps = []
        metric_values = []

        [row_num_timestamp, col_num] = coordinate_to_tuple(self.file_structure_details.metric_timestamp_seed)
        row_num_data = self.file_structure_details.metric_row

        for col_num in range(col_num, self.worksheet.max_column + 1):
            col_letter = get_column_letter(col_num)

            timestamp = self.worksheet[f"{col_letter}{row_num_timestamp}"].value
            if not isinstance(timestamp, datetime.datetime):
                timestamp = datetime.datetime.strptime(timestamp.strip(), self.file_structure_details.date_format)

            timestamps.append(timestamp)
            metric_values.append(self.worksheet[f"{col_letter}{row_num_data}"].value)

        return pd.Series(data=metric_values, index=timestamps).sort_index()

class MetricExtractorFileStyleA(IMetricExtractor):
    def get_company_name(self) -> str:
        return self.worksheet[self.file_structure_details.company_name_cell].value.split(" | ")[0].strip()


class MetricExtractorFileStyleB(IMetricExtractor):
    def get_company_name(self) -> str:
        return self.worksheet[self.file_structure_details.company_name_cell].value.split(" (")[0].strip()


@dataclass
class MetricOfCompany:
    company_name: str
    company_ticker: str
    metric_data: pd.Series

    def __repr__(self):
        data_status = "Data has been extracted." if not self.metric_data.empty else "Failed to extract data."
        return f"Company: {self.company_name}. {data_status}"

class PerMetricExtractor():
    def __init__(self, data_folder : str, file_styles_details : dict, file_style_to_metric_extractor_map : dict):
        self.data_folder = data_folder
        self.file_names = os.listdir(data_folder)
        self.styles_details = file_styles_details
        self.file_style_to_metric_extractor_map = file_style_to_metric_extractor_map

    def extract(self, data_frequency : FrequencyOfData):
        progress_bar = tqdm(total=int(len(self.file_names)/3)) # int() to dispaly x/int instead of x/float. Magic number 3 is justified by the fact that every company comes with 3 files ALWAYS.
        style_manager = FileStyleManager(self.styles_details)
        metrics_of_companies = []
        for file_name in self.file_names:
            [company_ticker, frequency, *_] = list(map(str.upper, file_name.split(".")[0].split("_"))) # Remove the extension of the file, get only the name of the company and the frequency of the data in caps and discard the rest

            if frequency != data_frequency.name:
                continue

            progress_bar.set_description(f"Processing {company_ticker}")
            progress_bar.update(1)

            workbook = opx.load_workbook(os.path.join(self.data_folder, file_name))
        
            file_style = style_manager.determine_file_style(workbook)
            extractor = self.determine_extractor(file_style)(workbook, self.styles_details[file_style])
            if not extractor:
                raise Exception("Unrecognized file style")

            company_name = extractor.get_company_name()
            metric_data = extractor.get_metric_data()

            metrics_of_companies.append(MetricOfCompany(company_name, company_ticker, metric_data))

        progress_bar.close()

        return metrics_of_companies

    def determine_extractor(self, file_style : FileStyle) -> IMetricExtractor:
        return self.file_style_to_metric_extractor_map.get(file_style, None) # Nice, method built into the dict


def main():
    # Set up objects
    extractor_classes = {
            FileStyle.A: MetricExtractorFileStyleA,
            FileStyle.B: MetricExtractorFileStyleB
        } # TODO: Go back to only having one class. The only difference between the classes is one string in the abstract method they share, add this string as a parameter in the FileStyleDetails dataclass
    # TODO: this could get a rework. telling the sheet and row name to go retrieve is redundant. An idea is to get the common row substring and search for it in every style to get the target row
    file_styles_details_ROA = {
            FileStyle.A: FileStyleDetails("A1", "%b-%Y", "Ratios - Key Metric", 28, "A1", "A3", "C6"),
            FileStyle.B: FileStyleDetails("B2", "%d-%m-%Y", "Financial Summary", 73, "A1", "A14", "B12")
        }
    ROA_frecuency = FrequencyOfData.QUARTERLY
    ROA_extractor = PerMetricExtractor("companies_data", file_styles_details_ROA, extractor_classes)

    # Execute main logic
    ROA = ROA_extractor.extract(ROA_frecuency)
    [print(roa) for roa in ROA]
    # Do stuff with data...

if __name__ == "__main__":
    main()

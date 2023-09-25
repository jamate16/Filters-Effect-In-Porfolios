from dataclasses import dataclass
from abc import ABC, abstractmethod
from enum import Enum

import openpyxl as opx
from openpyxl.utils import get_column_letter, column_index_from_string, coordinate_to_tuple

import os
import datetime
import numpy as np
import pandas as pd

from tqdm import tqdm

class FrequencyOfData(Enum):
    ANNUAL = 1
    QUARTERLY = 2

@dataclass
class FileStyleDetails:
    """
    Dataclass that stores the locations, names, and any other attribute that characterizes a sheet style
    
    Attributes:
        company_name_cell (str): The cell coordinate containing the company name.
        metric_sheet_name (str, optional): The name of the sheet where the metric is located (default is an empty string).
        metric_row_name (str, optional): The substring to search in the index of the rows of metric_sheet_name (default is an empty string).
        sheet_name_base (str): For comparison with sheet_name_compare. Cell coord where the sheet name is located.
        sheet_name_compare (str): For comparison with sheet_name_base. Cell coord where the sheet name is repeated.
    """
    company_name_cell: str
    company_name_separator: str
    date_format: str
    metric_sheet_name: str
    metric_name: str
    sheet_name_base: str
    sheet_name_compare: str
    metric_timestamp_seed: str

class FileStyle(Enum):
    A = "StyleA"
    B = "StyleB"

class FileStyleManager:
    def __init__(self, styles : dict):
        self.styles = styles

    def determine_file_style(self, workbook : opx.Workbook) -> FileStyle:
        """
        Compares the contents of two cells (different cells for every file style) that contain the same substring

        Returns:
            FileStyle: If the comparison for the corresponding file style returns true
        """
        active_sheet = workbook.active
        
        for style_name, style in self.styles.items():
            sheet_name_base = active_sheet[style.sheet_name_base].value
            sheet_name_compare = active_sheet[style.sheet_name_compare].value.split("-") #
            sheet_name_compare = sheet_name_compare[0].split("\xa0\xa0")[0].strip() # In style A there a bunch of "\xa0" characters

            if sheet_name_compare in sheet_name_base:
                return style_name
            
        raise Exception("Sheet style not recognized.")

class MetricNotFoundInSheet(Exception):
    def __init__(self, metric_name: str, sheet_name: str):
        self.metric_name = metric_name
        self.sheet_name = sheet_name

    def __str__(self):
        return(repr(f"Metric {self.metric_name} not found in {self.sheet_name}."))

class IMetricExtractor(ABC):
    def __init__(self, workbook : opx.Workbook, file_structure_details : FileStyleDetails, quarter_end_dates : list = ["31-03", "30-06", "30-09", "31-12"]):
        self.workbook = workbook
        self.file_structure_details = file_structure_details

        self.worksheet = None
        self.open_sheet()

        self.quarter_end_dates = quarter_end_dates

    def open_sheet(self) -> None:
        for sheet in self.workbook.worksheets:
            if self.file_structure_details.metric_sheet_name in sheet[self.file_structure_details.sheet_name_base].value:
                self.worksheet = sheet
                break

        if not self.worksheet:
            raise Exception(f"Metric sheet not found. {self.file_structure_details.metric_sheet_name} not in cell {self.file_structure_details.sheet_name_base} on any sheet")

    def get_company_name(self) -> str:
        return self.worksheet[self.file_structure_details.company_name_cell].value.split(self.file_structure_details.company_name_separator)[0].strip()

    def format_date(self, date: datetime.datetime, quarter: int, year: str):
        full_date_str = self.quarter_end_dates[quarter-1] + "-" + year # Offset of -1: quarters 1-4, index 0-3

        return datetime.datetime.strptime(full_date_str, "%d-%m-%Y")

    def get_metric_data(self) -> pd.Series:
        timestamps = []
        metric_values = []

        # Get the col and rows where the extraction starts
        [row_num_timestamp, col_num] = coordinate_to_tuple(self.file_structure_details.metric_timestamp_seed)

        # Find row that contains the target metric
        row_num_data = self.find_row_with_target_metric()

        # Variables for date formatting
        quarters = 4
        current_year = None
        quarters_till_next_year = 0

        for col_num in range(col_num, self.worksheet.max_column + 1):
            col_letter = get_column_letter(col_num)

            timestamp = self.worksheet[f"{col_letter}{row_num_timestamp}"].value
            if not isinstance(timestamp, datetime.datetime):
                timestamp = datetime.datetime.strptime(timestamp.strip(), self.file_structure_details.date_format)

            # Determine the quarter
            year_cell_value = self.worksheet[f"{get_column_letter(col_num)}{row_num_timestamp}"].value.strip()[-4:]
            if year_cell_value != current_year and year_cell_value is not None: # Every time I am on a new year, count how many quarters until the next
                current_year = year_cell_value
                quarters_till_next_year = 0
                # Calculate quarters until next year by comparing the current value against the next 4 which would yield 4 quarters until next year
                for i in range(1, quarters+1): 
                    year_cell_value_i = self.worksheet[f"{get_column_letter(col_num+i)}{row_num_timestamp-1}"].value
                    quarters_till_next_year += 1
                    if not (year_cell_value_i == current_year or year_cell_value_i is None):
                        break

            quarter = self.calculate_fiscal_quarter(quarters_till_next_year, quarters)

            timestamp = self.format_date(timestamp, quarter, current_year.strip())
            quarters_till_next_year -= 1

            timestamps.append(timestamp)
            metric_cell = self.worksheet[f"{col_letter}{row_num_data}"]
            if metric_cell.value == None:
                metric_values.append(np.NaN)
            else:
                if metric_cell.number_format == "[>=100]##,##0.0\%;[<=-100]\-##,##0.0\%;##,##0.0\%":
                    metric_values.append(metric_cell.value/100)
                else:
                    metric_values.append(metric_cell.value)
                

        return pd.Series(data=metric_values, index=timestamps).sort_index()

    def find_row_with_target_metric(self) -> int:
        row_num_data = 0
        for cell in self.worksheet['A']:
            if cell.value is not None and self.file_structure_details.metric_name in cell.value:
                row_num_data = cell.row
                break

        if row_num_data == 0:
            raise MetricNotFoundInSheet(self.file_structure_details.metric_name, self.file_structure_details.metric_sheet_name)

        return row_num_data

    @abstractmethod
    def calculate_fiscal_quarter(self, qs_till_next_year: int, qs_in_year):
        pass

# TODO: Assess if we need strategy pattern here. Lots or repeated code in the concrete implementations of the interface.
class MetricExtractorFileStyleA(IMetricExtractor):
    def calculate_fiscal_quarter(self, qs_till_next_year: int, qs_in_year):
        return qs_till_next_year # Date is descending order

class MetricExtractorFileStyleB(IMetricExtractor):
    def calculate_fiscal_quarter(self, qs_till_next_year: int, qs_in_year):
        return qs_in_year - (qs_till_next_year - 1) # If qtny is one, it means we are in q4, so current_quarter = 4 - (1-1) which returns quarter 4 as expected

@dataclass
class MetricOfCompany:
    company_name: str
    company_ticker: str
    metric_data: pd.Series

    def __repr__(self):
        data_status = "Data has been extracted." if not self.metric_data.empty else "Failed to extract data."
        return f"Company: {self.company_name}. {data_status}"

class MetricExtractor():
    def __init__(self, data_folder : str, file_styles_details : dict, file_style_to_metric_extractor_map : dict):
        self.data_folder_path = os.path.join(os.path.dirname(__file__), "..", data_folder)

        self.file_names = os.listdir(self.data_folder_path)
        self.styles_details = file_styles_details
        self.file_style_to_metric_extractor_map = file_style_to_metric_extractor_map

        self.extracted_data = None

        # Status of last extraction
        self.companies_successfully_extracted = 0
        self.companies_with_not_enough_data = []

    def extract(self, data_frequency : FrequencyOfData):
        progress_bar = tqdm(total=int(len(self.file_names)/2)) # int() to dispaly x/int instead of x/float. Magic number 2 represents that almost all compnies only have 2 files
        style_manager = FileStyleManager(self.styles_details)
        
        metrics_of_companies = []
        for file_name in self.file_names:
            # Remove the extension of the file, get only the name of the company and the frequency of the data in caps and discard the rest
            [company_ticker, frequency, *_] = list(map(str.upper, file_name.split(".")[0].split("_")))

            if frequency != data_frequency.name:
                continue

            progress_bar.set_description(f"Processing {company_ticker}")
            progress_bar.update(1)

            workbook = opx.load_workbook(os.path.join(self.data_folder_path, file_name))

            # Workbooks with less than 4 sheets are worthless to us
            if (len(workbook.worksheets) < 4):
                self.companies_with_not_enough_data.append(company_ticker)
                continue
        
            file_style = style_manager.determine_file_style(workbook)
            extractor = self.determine_extractor(file_style)(workbook, self.styles_details[file_style])
            if not extractor:
                raise Exception("Unrecognized file style")

            company_name = extractor.get_company_name()
            try:
                metric_data = extractor.get_metric_data().rename(company_ticker) # Give the pd.Series a name, this will later be the name of the col
            except MetricNotFoundInSheet as e:
                print(str(e))
                self.companies_with_not_enough_data.append(company_ticker)
                continue

            metrics_of_companies.append(MetricOfCompany(company_name, company_ticker, metric_data))

        progress_bar.close()

        self.companies_successfully_extracted = len(metrics_of_companies)
        self.print_extraction_summary()
        
        self.extracted_data = metrics_of_companies

    def determine_extractor(self, file_style : FileStyle) -> IMetricExtractor:
        
        
        return self.file_style_to_metric_extractor_map.get(file_style, None) # Nice, method built into the dict

    def print_extraction_summary(self) -> None:
        summary_str = f"Successfully extracted data for {self.companies_successfully_extracted}. "
        summary_str += f"{self.companies_with_not_enough_data} companies ignored, corresponding workbook incomplete or metric not found in target sheet." if len(self.companies_with_not_enough_data) > 0 else ""
        print(summary_str)

    def get_dataframe(self) -> pd.DataFrame:
        merged = pd.DataFrame()

        for metric_object in self.extracted_data:
            merged = pd.merge(merged, metric_object.metric_data, how='outer', left_index=True, right_index=True)
        
        return merged

def main():
    # Set up objects
    extractor_classes = {
            FileStyle.A: MetricExtractorFileStyleA,
            FileStyle.B: MetricExtractorFileStyleB
        } # TODO: Go back to only having one class. The only difference between the classes is one string in the abstract method they share, add this string as a parameter in the FileStyleDetails dataclass
    # TODO: this could get a rework. telling the sheet and row name to go retrieve is redundant. An idea is to get the common row substring and search for it in every style to get the target row
    file_styles_details_ROA = {
            FileStyle.A: FileStyleDetails("A1", " | ", "%b-%Y", "Ratios - Key Metric", "Pretax ROA", "A1", "A3", "C6"),
            FileStyle.B: FileStyleDetails("B2", " (", "%d-%m-%Y", "Financial Summary", "Pretax ROA", "A1", "A14", "B15")
        }
    ROA_frecuency = FrequencyOfData.QUARTERLY
    ROA_extractor = MetricExtractor("companies_data", file_styles_details_ROA, extractor_classes)

    # Execute main logic
    ROA_extractor.extract(ROA_frecuency)
    ROA = ROA_extractor.extracted_data
    df = ROA_extractor.get_dataframe()

if __name__ == "__main__":
    main()

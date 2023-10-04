from dataclasses import dataclass
from abc import ABC, abstractmethod
from enum import Enum
import pickle

import openpyxl as opx
from openpyxl.utils import get_column_letter, column_index_from_string, coordinate_to_tuple

import os
import datetime
import numpy as np
import pandas as pd

from tqdm import tqdm

from .file_style import FileStyle, FileStyleDetails, FileStyleManager

class FrequencyOfData(Enum):
    ANNUAL = 1
    QUARTERLY = 2

class MetricNotFoundInSheet(Exception):
    def __init__(self, metric_name: str, sheet_name: str):
        self.metric_name = metric_name
        self.sheet_name = sheet_name

    def __str__(self):
        return(repr(f"Metric {self.metric_name} not found in {self.sheet_name}."))

class IMetricFetcher(ABC):
    def __init__(self, workbook : opx.Workbook, file_structure_details : FileStyleDetails, quarter_end_dates : list = ["31-03", "30-06", "30-09", "31-12"]):
        self.workbook = workbook
        self.file_structure_details = file_structure_details

        self.worksheet = None
        self.open_sheet()

        self.quarter_end_dates = quarter_end_dates

    def open_sheet(self) -> None:
        for sheet in self.workbook.worksheets:
            if self.file_structure_details.metric_sheet_name in sheet[self.file_structure_details.sheet_name_base].value:
                self.worksheet = sheet
                break

        if not self.worksheet:
            raise Exception(f"Metric sheet not found. {self.file_structure_details.metric_sheet_name} not in cell {self.file_structure_details.sheet_name_base} on any sheet")

    def get_company_name(self) -> str:
        return self.worksheet[self.file_structure_details.company_name_cell].value.split(self.file_structure_details.company_name_separator)[0].strip()

    def format_date(self, quarter: int, year: str):
        full_date_str = self.quarter_end_dates[quarter-1] + "-" + year # Offset of -1: quarters 1-4, index 0-3

        return datetime.datetime.strptime(full_date_str, "%d-%m-%Y").date()

    def get_metric_data(self) -> pd.Series:
        timestamps = []
        metric_values = []

        # Get the col and rows where the extraction starts
        [row_num_timestamp, col_num] = coordinate_to_tuple(self.file_structure_details.timestamp_coord)

        # Find row that contains the target metric
        row_num_data = self.find_row_with_target_metric()

        # Variables for date formatting
        quarters = 4
        current_year = None
        quarters_till_next_year = 0

        for col_num in range(col_num, self.worksheet.max_column + 1):
            col_letter = get_column_letter(col_num)

            # Determine the quarter
            year_cell_value = self.worksheet[f"{get_column_letter(col_num)}{row_num_timestamp}"].value
            if year_cell_value != current_year and year_cell_value is not None: # Every time I am on a new year, count how many quarters until the next
                current_year = year_cell_value
                quarters_till_next_year = 0
                # Calculate quarters until next year by comparing the current value against the next 4 which would yield 4 quarters until next year
                for i in range(1, quarters+1): 
                    year_cell_value_i = self.worksheet[f"{get_column_letter(col_num+i)}{row_num_timestamp}"].value
                    quarters_till_next_year += 1
                    if not (year_cell_value_i == current_year or year_cell_value_i is None):
                        break

            quarter = self.calculate_fiscal_quarter(quarters_till_next_year, quarters)
            timestamp = self.format_date(quarter, current_year.strip())
            quarters_till_next_year -= 1

            timestamps.append(timestamp)
            metric_cell = self.worksheet[f"{col_letter}{row_num_data}"]
            if metric_cell.value == None:
                metric_values.append(np.NaN)
            else:
                if metric_cell.number_format == r"[>=100]##,##0.0\%;[<=-100]\-##,##0.0\%;##,##0.0\%":
                    metric_values.append(metric_cell.value/100)
                else:
                    metric_values.append(metric_cell.value)
                
        return pd.Series(data=metric_values, index=timestamps)

    def find_row_with_target_metric(self) -> int:
        row_num_data = 0
        for cell in self.worksheet['A']:
            if cell.value is not None and self.file_structure_details.metric_name in cell.value:
                row_num_data = cell.row
                break

        if row_num_data == 0:
            raise MetricNotFoundInSheet(self.file_structure_details.metric_name, self.file_structure_details.metric_sheet_name)

        return row_num_data

    @abstractmethod
    def calculate_fiscal_quarter(self, qs_till_next_year: int, qs_in_year):
        pass

# TODO: Assess if we need strategy pattern here. Lots or repeated code in the concrete implementations of the interface.
class MetricFetcherFileStyleA(IMetricFetcher):
    def calculate_fiscal_quarter(self, qs_till_next_year: int, qs_in_year):
        return qs_till_next_year # Date is descending order

class MetricFetcherFileStyleB(IMetricFetcher):
    def calculate_fiscal_quarter(self, qs_till_next_year: int, qs_in_year):
        return qs_in_year - (qs_till_next_year - 1) # If qtny is one, it means we are in q4, so current_quarter = 4 - (1-1) which returns quarter 4 as expected

class MetricFetcherFileStyleFactory:
    @staticmethod
    def get_extractor(file_style: FileStyle, *extractor_args):
        match file_style:
            case FileStyle.A:
                return MetricFetcherFileStyleA(*extractor_args)
            case FileStyle.B:
                return MetricFetcherFileStyleB(*extractor_args)
            case _:
                return None

@dataclass
class MetricOfCompany:
    company_name: str
    company_ticker: str
    metric_data: pd.Series

def __repr__(self):
        data_status = "Data has been extracted." if not self.metric_data.empty else "Failed to extract data."
        return f"Company: {self.company_name}. {data_status}"

class MetricsFetcher:
    def __init__(self, data_folder_path : str, file_style_configs_by_metrics : dict):
        self.data_folder_path = data_folder_path
        self.file_names = os.listdir(self.data_folder_path)

        self.file_style_configs_by_metrics = file_style_configs_by_metrics

        # Status of last extraction
        self.extracted_data = None
        self.companies_successfully_extracted = 0
        self.companies_with_not_enough_data = []

    def _load_from_pickle_file(self, full_file_path) -> pd.DataFrame:
        with open(full_file_path, "rb") as infile:
            metric_df = pickle.load(infile)
            return metric_df
    
    def _load_from_excel_file(self, metric: str, data_frequency: FrequencyOfData=FrequencyOfData.QUARTERLY):
        file_style_configs = self.file_style_configs_by_metrics[metric]
        style_manager = FileStyleManager(file_style_configs)
        
        progress_bar = tqdm(total=int(len(self.file_names)/2), position=0, leave=True) # int() to dispaly x/int instead of x/float. Magic number 2 represents that almost all compnies only have 2 files
        metrics_of_companies = []
        for file_name in self.file_names:
            # Remove the extension of the file, get only the name of the company and the frequency of the data in caps and discard the rest
            [company_ticker, frequency, *_] = list(map(str.upper, file_name.split(".")[0].split("_")))

            if frequency != data_frequency.name:
                continue

            progress_bar.set_description(f"Processing {company_ticker}")
            progress_bar.update(1)

            workbook = opx.load_workbook(os.path.join(self.data_folder_path, file_name))

            # Workbooks with less than 4 sheets are worthless to us
            if (len(workbook.worksheets) < 4):
                self.companies_with_not_enough_data.append(company_ticker)
                continue
        
            file_style = style_manager.determine_file_style(workbook)
            extractor = MetricFetcherFileStyleFactory.get_extractor(file_style, workbook, file_style_configs[file_style])
            
            if not extractor:
                raise Exception("Unrecognized file style")

            company_name = extractor.get_company_name()
            try:
                metric_data = extractor.get_metric_data().rename(company_ticker) # Give the pd.Series a name, this will later be the name of the col
            except MetricNotFoundInSheet as e:
                # print(str(e))
                self.companies_with_not_enough_data.append(company_ticker)
                continue

            metrics_of_companies.append(MetricOfCompany(company_name, company_ticker, metric_data))

        progress_bar.close()

        self.companies_successfully_extracted = len(metrics_of_companies)
        self.print_extraction_summary()
        
        self.extracted_data = metrics_of_companies.copy()

    def _save_metric_data(self, full_file_path: str, data: pd.DataFrame):
        with open(full_file_path, "wb") as outfile:
            pickle.dump(data, outfile)

    def fetch(self,
              metric: str,
              pickled_data_path: str=os.path.join("..", "..", "data", "pickled_data"),
              data_frequency: FrequencyOfData=FrequencyOfData.QUARTERLY):
        
        pickled_data_file_path = os.path.join(pickled_data_path, f"{metric}_data.pickle") # TODO: move this configuration to a file in its corresponding folder inside or src
        try:
            metric_df = self._load_from_pickle_file(pickled_data_file_path)
        except FileNotFoundError:
            self._load_from_excel_file(metric, data_frequency)
            metric_df = self.get_dataframe()
        # TODO: Remove the need to pass the data to _save_metric_data by, perhaps, saving self.extracted_data instead of what returns self.get_dataframe()
        self._save_metric_data(pickled_data_file_path, metric_df)
        return metric_df

    def print_extraction_summary(self) -> None:
        summary_str = f"Successfully extracted data for {self.companies_successfully_extracted}. "
        summary_str += f"{self.companies_with_not_enough_data} companies ignored, corresponding workbook incomplete or metric not found in target sheet." if len(self.companies_with_not_enough_data) > 0 else ""
        print(summary_str)

    def get_dataframe(self) -> pd.DataFrame:
        merged = pd.DataFrame()

        for metric_object in self.extracted_data:
            merged = pd.merge(merged, metric_object.metric_data, how='outer', left_index=True, right_index=True)
        
        return merged.copy()

def main():
    # extractor = MetricsExtractor("companies_data", file_style_configs_by_metric)

    # # Execute main logic
    # extractor.extract("Pretax ROA")
    # ROA = extractor.extracted_data
    # df = extractor.get_dataframe()
    pass

if __name__ == "__main__":
    main()
